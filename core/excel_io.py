"""엑셀 입출력.

입력: A열=업체명. 헤더 자동감지 + 지역/업종 힌트 컬럼 자동 매핑.
출력: 원본 시트 보존 + 우측에 결과 컬럼 추가. 별도 `_후보업체` 시트에 동명 후보 덤프.
"""
from __future__ import annotations

import copy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# 결과 컬럼 순서 (그대로 메인 시트 우측에 추가됨)
RESULT_COLUMNS = [
    "매칭상태",
    "매칭된업체명",
    "ICP점수",
    "대표번호",
    "신뢰도",
    "출처",
    "주소_시도",
    "주소_시군구",
    "주소_동",
    "주소_전체",
    "후보번호",
    "비고",
]

CANDIDATE_SHEET_NAME = "_후보업체"
CANDIDATE_HEADERS = ["원본업체명", "후보순위", "네이버명칭", "카테고리", "주소", "전화", "ICP점수"]

# 지역/업종 힌트로 인식할 헤더 키워드
_REGION_HINT_TOKENS = ("주소", "address", "소재지", "지역", "시도", "광역시", "시군구", "구")
_CATEGORY_HINT_TOKENS = ("업종", "카테고리", "분야", "산업", "category")


@dataclass
class InputRow:
    row_index: int           # openpyxl 행 번호 (1-base)
    company_name: str
    region_hint: str
    category_hint: str


@dataclass
class ExcelInput:
    workbook_bytes: bytes    # 원본 그대로 복사
    sheet_name: str
    header_row: int          # 0이면 헤더 없음, 1+ 이면 그 행이 헤더
    company_col: int         # 1-base (A=1)
    region_hint_col: int     # 0이면 없음
    category_hint_col: int   # 0이면 없음
    rows: list[InputRow]


def parse_input(file_bytes: bytes) -> ExcelInput:
    """업로드된 엑셀 바이트를 파싱."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    sheet_name = ws.title

    # 헤더 행 자동 감지: 1행이 모두 문자열이고 2행이 일반 데이터면 헤더.
    header_row = _detect_header_row(ws)
    region_col, category_col = _detect_hint_columns(ws, header_row)

    rows: list[InputRow] = []
    start = (header_row + 1) if header_row else 1
    for r in range(start, ws.max_row + 1):
        name = _cell_str(ws.cell(row=r, column=1).value)
        if not name:
            continue
        region_hint = _cell_str(ws.cell(row=r, column=region_col).value) if region_col else ""
        category_hint = _cell_str(ws.cell(row=r, column=category_col).value) if category_col else ""
        rows.append(InputRow(
            row_index=r,
            company_name=name,
            region_hint=region_hint,
            category_hint=category_hint,
        ))

    return ExcelInput(
        workbook_bytes=file_bytes,
        sheet_name=sheet_name,
        header_row=header_row,
        company_col=1,
        region_hint_col=region_col,
        category_hint_col=category_col,
        rows=rows,
    )


def write_output(
    excel_input: ExcelInput,
    results_by_row: dict[int, dict],     # row_index -> result dict
    candidates_by_row: dict[int, list[dict]],   # row_index -> [{rank, title, category, address, phone, icp_score}]
) -> bytes:
    """원본 워크북에 결과 컬럼을 추가해 새 워크북 바이트 반환."""
    wb = load_workbook(BytesIO(excel_input.workbook_bytes))
    ws = wb[excel_input.sheet_name]

    # 결과 컬럼 시작 위치 (기존 max_column 다음)
    start_col = ws.max_column + 1

    # 헤더 작성
    header_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    bold = Font(bold=True)
    header_row_idx = excel_input.header_row if excel_input.header_row else 1
    for i, name in enumerate(RESULT_COLUMNS):
        cell = ws.cell(row=header_row_idx, column=start_col + i, value=name)
        cell.font = bold
        cell.fill = header_fill

    # 본문 작성
    for row in excel_input.rows:
        result = results_by_row.get(row.row_index) or {}
        values = [
            result.get("매칭상태", ""),
            result.get("매칭된업체명", ""),
            result.get("ICP점수", ""),
            result.get("대표번호", ""),
            result.get("신뢰도", ""),
            result.get("출처", ""),
            result.get("주소_시도", ""),
            result.get("주소_시군구", ""),
            result.get("주소_동", ""),
            result.get("주소_전체", ""),
            result.get("후보번호", ""),
            result.get("비고", ""),
        ]
        for i, v in enumerate(values):
            ws.cell(row=row.row_index, column=start_col + i, value=v)

    # 후보 시트
    if any(candidates_by_row.values()):
        if CANDIDATE_SHEET_NAME in wb.sheetnames:
            del wb[CANDIDATE_SHEET_NAME]
        cs = wb.create_sheet(CANDIDATE_SHEET_NAME)
        for i, h in enumerate(CANDIDATE_HEADERS, start=1):
            c = cs.cell(row=1, column=i, value=h)
            c.font = bold
            c.fill = header_fill
        write_row = 2
        for row in excel_input.rows:
            cands = candidates_by_row.get(row.row_index) or []
            for cand in cands:
                cs.cell(row=write_row, column=1, value=row.company_name)
                cs.cell(row=write_row, column=2, value=cand.get("rank", ""))
                cs.cell(row=write_row, column=3, value=cand.get("title", ""))
                cs.cell(row=write_row, column=4, value=cand.get("category", ""))
                cs.cell(row=write_row, column=5, value=cand.get("address", ""))
                cs.cell(row=write_row, column=6, value=cand.get("phone", ""))
                cs.cell(row=write_row, column=7, value=cand.get("icp_score", ""))
                write_row += 1
        # 컬럼 폭 조정
        widths = [16, 8, 24, 24, 36, 16, 10]
        for i, w in enumerate(widths, start=1):
            cs.column_dimensions[get_column_letter(i)].width = w

    # 메인 시트 결과 컬럼 폭 조정
    width_map = {0: 12, 1: 22, 2: 8, 3: 16, 4: 10, 5: 24, 6: 14, 7: 18, 8: 12, 9: 32, 10: 24, 11: 28}
    for i, w in width_map.items():
        ws.column_dimensions[get_column_letter(start_col + i)].width = w

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _detect_header_row(ws) -> int:
    """1행에 명시적인 헤더 키워드가 있을 때만 헤더로 판정.

    이름만으로는 헤더인지 데이터인지 구분이 어려운 경우(예: A열에 "삼성전자"만 있는 시트)
    오인해서 첫 행을 통째로 누락시키는 사고를 막기 위해 보수적으로 동작.
    """
    if ws.max_row < 1:
        return 0
    row1 = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    first = row1[0] if row1 else None
    header_tokens = ("업체", "회사", "기업", "상호", "법인", "name", "Name", "NAME", "거래처")
    if isinstance(first, str) and any(k in first for k in header_tokens):
        return 1
    # 다른 컬럼에 힌트 키워드(주소·업종 등)가 있으면 헤더로 판정
    for v in row1[1:]:
        if not isinstance(v, str):
            continue
        h = v.strip()
        if any(tok in h for tok in _REGION_HINT_TOKENS) or any(tok in h for tok in _CATEGORY_HINT_TOKENS):
            return 1
    return 0


def _detect_hint_columns(ws, header_row: int) -> tuple[int, int]:
    if not header_row:
        return 0, 0
    region_col = 0
    category_col = 0
    for c in range(2, ws.max_column + 1):  # A열은 업체명이므로 B열부터
        v = ws.cell(row=header_row, column=c).value
        if not isinstance(v, str):
            continue
        header = v.strip()
        if not region_col and any(tok in header for tok in _REGION_HINT_TOKENS):
            region_col = c
            continue
        if not category_col and any(tok in header for tok in _CATEGORY_HINT_TOKENS):
            category_col = c
            continue
    return region_col, category_col


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_sample_workbook() -> bytes:
    """샘플 입력 엑셀(15개 업체) 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = "업체리스트"
    ws["A1"] = "업체명"
    ws["B1"] = "지역"
    ws["C1"] = "업종"

    samples = [
        # (업체명, 지역힌트, 업종힌트)
        ("현대건설", "서울", "건설"),
        ("대우건설", "서울", "건설"),
        ("CJ대한통운", "서울", "물류"),
        ("한국전력공사", "전남 나주", "공기업"),
        ("포스코", "포항", "제조"),
        ("두산중공업", "창원", "중공업"),
        ("삼성중공업", "거제", "조선"),
        ("KT", "서울", "통신"),
        ("한국가스공사", "대구", "공기업"),
        ("한국도로공사", "김천", "공기업"),
        ("한국건설", "", ""),                # 동명 후보 다수
        ("동방운수", "", ""),                # 동명 후보 다수
        ("동양물산", "", "농기계"),
        ("이름없는가상회사123", "", ""),     # 검색 실패 예상
        ("존재안함XYZ주식회사", "", ""),     # 검색 실패 예상
    ]
    for i, (name, region, category) in enumerate(samples, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=region)
        ws.cell(row=i, column=3, value=category)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    for c in (1, 2, 3):
        ws.cell(row=1, column=c).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
