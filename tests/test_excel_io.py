from io import BytesIO

from openpyxl import Workbook, load_workbook

from core.excel_io import build_sample_workbook, parse_input, write_output


def _make_workbook(rows, headers=None):
    wb = Workbook()
    ws = wb.active
    r = 1
    if headers:
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r, column=c, value=h)
        r += 1
    for row in rows:
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_with_headers_and_hints():
    data = _make_workbook(
        rows=[
            ["현대건설", "서울 강남구", "건설"],
            ["대우건설", "서울 종로구", "건설"],
        ],
        headers=["업체명", "주소", "업종"],
    )
    parsed = parse_input(data)
    assert parsed.header_row == 1
    assert parsed.region_hint_col == 2
    assert parsed.category_hint_col == 3
    assert len(parsed.rows) == 2
    assert parsed.rows[0].company_name == "현대건설"
    assert parsed.rows[0].region_hint == "서울 강남구"
    assert parsed.rows[0].category_hint == "건설"


def test_parse_without_headers():
    data = _make_workbook(
        rows=[["삼성전자"], ["LG전자"]],
    )
    parsed = parse_input(data)
    assert parsed.header_row == 0
    assert parsed.rows[0].company_name == "삼성전자"
    assert parsed.region_hint_col == 0
    assert parsed.category_hint_col == 0


def test_write_output_adds_columns():
    data = _make_workbook(
        rows=[["현대건설", "서울", "건설"]],
        headers=["업체명", "주소", "업종"],
    )
    parsed = parse_input(data)
    row_index = parsed.rows[0].row_index
    results = {
        row_index: {
            "매칭상태": "매칭확정",
            "매칭된업체명": "현대건설(주)",
            "ICP점수": 5.0,
            "대표번호": "02-1234-5678",
            "신뢰도": "검증됨",
            "출처": "지도+홈페이지",
            "주소_시도": "서울특별시",
            "주소_시군구": "종로구",
            "주소_동": "계동",
            "주소_전체": "서울특별시 종로구 율곡로 75",
            "후보번호": "02-1234-5678",
            "비고": "",
        }
    }
    out = write_output(parsed, results, {})
    wb = load_workbook(BytesIO(out))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "매칭상태" in headers
    assert "주소_동" in headers
    assert "_후보업체" not in wb.sheetnames  # 확정 케이스만 있으므로 후보 시트 생략

    # 결과 값이 제대로 박혔는지
    last_col = ws.max_column
    assert ws.cell(row=2, column=last_col).value == "" or ws.cell(row=2, column=last_col).value is None or ws.cell(row=2, column=last_col).value == ""


def test_write_output_with_candidate_sheet():
    data = _make_workbook(rows=[["한국건설"]])
    parsed = parse_input(data)
    row_index = parsed.rows[0].row_index
    results = {row_index: {"매칭상태": "확정필요", "비고": "동명 후보 다수"}}
    cands = {row_index: [
        {"rank": 1, "title": "한국건설 본사", "category": "종합건설", "address": "서울특별시 강남구 역삼동", "phone": "02-1234-5678", "icp_score": 5},
        {"rank": 2, "title": "한국건설식당", "category": "한식", "address": "서울특별시 마포구 합정동", "phone": "02-9999-9999", "icp_score": -3},
    ]}
    out = write_output(parsed, results, cands)
    wb = load_workbook(BytesIO(out))
    assert "_후보업체" in wb.sheetnames
    cs = wb["_후보업체"]
    assert cs.cell(row=1, column=1).value == "원본업체명"
    assert cs.cell(row=2, column=3).value == "한국건설 본사"


def test_sample_workbook_round_trip():
    data = build_sample_workbook()
    parsed = parse_input(data)
    assert len(parsed.rows) >= 10
    assert parsed.region_hint_col != 0
    assert parsed.category_hint_col != 0
